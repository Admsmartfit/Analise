import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="2C5F5B")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells), default=0)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 45)


def _plano_label(plano):
    return "Black+" if plano == "black_plus" else plano.capitalize()


def _venda_label(key):
    """Converte a chave interna 'venda_detalhe|canal|plano|periodo' em um rótulo legível."""
    parts = key.split('|')
    if len(parts) == 4:
        _, canal, plano, periodo = parts
        return f"Vendas {canal.capitalize()} {_plano_label(plano)} | {periodo.capitalize()}"
    return key


def _pre_venda_label(key):
    """Converte a chave interna 'pre_venda_detalhe|canal|plano' em um rótulo legível."""
    parts = key.split('|')
    if len(parts) == 3:
        _, canal, plano = parts
        return f"Vendas {canal.capitalize()} {_plano_label(plano)}"
    return key


def _cancel_label(key):
    parts = key.split('|')
    if len(parts) == 2:
        return f"Cancelados {parts[1].capitalize()}"
    return key


def export_to_xlsx(conn, output_path, data_referencia=None):
    """Exporta todos os dados capturados (formato padrão + pré-vendas) para uma planilha .xlsx,
    reconstituindo as colunas originais do e-mail para facilitar a conferência manual.
    """
    where_main = "WHERE f.data_referencia = %s" if data_referencia else ""
    where_pv = "WHERE f.data_referencia = %s" if data_referencia else ""
    params = (data_referencia,) if data_referencia else ()

    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT f.data_referencia, f.sigla_no_dia, u.nome_digital, u.pais, f.bloco_email, u.tipo_operacao,
                   u.data_inauguracao, f.unidade_imatura,
                   f.total_ativos, f.ativos_smart, f.ativos_black, f.ativos_fit, f.ativos_black_plus, f.ativos_studio, f.bloqueados,
                   f.visitas_dia, f.visitas_mes, f.conversao_dia, f.conversao_mes,
                   f.vendas_geral_dia, f.vendas_geral_mes, f.transferencias_liquida_mes,
                   f.detalhe_vendas_json, f.detalhe_movimentacoes_json, f.gmail_message_id
            FROM fato_metricas_diarias f
            JOIN dim_unidade u ON u.id = f.unidade_id
            {where_main}
            ORDER BY f.bloco_email, u.nome_digital;
        """, params)
        main_rows = cur.fetchall()

        cur.execute(f"""
            SELECT f.data_referencia, u.codigo_unidade, u.nome_unidade, u.pais_regiao,
                   f.vendas_total, f.detalhe_vendas_json, f.gmail_message_id
            FROM fato_pre_venda_diaria f
            JOIN dim_unidade_pre_venda u ON u.id = f.unidade_pre_venda_id
            {where_pv}
            ORDER BY u.pais_regiao, u.nome_unidade;
        """, params)
        pv_rows = cur.fetchall()

    # Descobre dinamicamente todas as combinações canal/plano/período já capturadas
    venda_keys = set()
    cancel_keys = set()
    for row in main_rows:
        venda_keys.update((row[22] or {}).keys())
        cancel_keys.update(((row[23] or {}).get('cancelamentos') or {}).keys())
    venda_keys = sorted(venda_keys)
    cancel_keys = sorted(cancel_keys)

    pv_keys = set()
    for row in pv_rows:
        pv_keys.update((row[5] or {}).keys())
    pv_keys = sorted(pv_keys)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- Aba Resumo ---
    ws_resumo = wb.create_sheet("Resumo")
    _write_header(ws_resumo, ["Métrica", "Valor"])
    ws_resumo.append(["Unidades — formato padrão", len(main_rows)])
    ws_resumo.append(["Unidades — pré-venda", len(pv_rows)])
    ws_resumo.append(["Filtro de data aplicado", str(data_referencia) if data_referencia else "Todas as datas"])
    ws_resumo.append(["Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
    _autosize(ws_resumo)

    # --- Aba Unidades (formato padrão) ---
    ws = wb.create_sheet("Unidades")
    headers = [
        "Data Referência", "Sigla", "Nome", "País", "Região/Bloco do E-mail", "Tipo Operação",
        "Inauguração", "Imatura?",
        "Ativos Total", "Ativos Smart", "Ativos Black", "Ativos Fit", "Ativos Black+", "Ativos Studio", "Ativos Bloqueados",
        "Visitas Dia", "Visitas Mês", "Conversão Dia (%)", "Conversão Mês (%)",
    ] + [_venda_label(k) for k in venda_keys] + [
        "Vendas Total Dia", "Vendas Total Mês", "Transferências Mês",
    ] + [_cancel_label(k) for k in cancel_keys] + [
        "ID do E-mail (UID)"
    ]
    _write_header(ws, headers)

    for row in main_rows:
        (data_ref, sigla, nome, pais, regiao, tipo_op, data_inaug, imatura,
         total_ativos, ativos_smart, ativos_black, ativos_fit, ativos_bp, ativos_studio, bloqueados,
         visitas_dia, visitas_mes, conv_dia, conv_mes,
         vendas_dia, vendas_mes, transferencias,
         detalhe_vendas, detalhe_mov, msg_id) = row

        detalhe_vendas = detalhe_vendas or {}
        cancelamentos = (detalhe_mov or {}).get('cancelamentos') or {}

        line = [
            data_ref, sigla, nome, pais, regiao, tipo_op, data_inaug, "Sim" if imatura else "Não",
            total_ativos, ativos_smart, ativos_black, ativos_fit, ativos_bp, ativos_studio, bloqueados,
            visitas_dia, visitas_mes, conv_dia, conv_mes,
        ] + [detalhe_vendas.get(k, 0) for k in venda_keys] + [
            vendas_dia, vendas_mes, transferencias,
        ] + [cancelamentos.get(k, 0) for k in cancel_keys] + [
            msg_id
        ]
        ws.append(line)

    _autosize(ws)
    ws.freeze_panes = "D2"

    # --- Aba Pré-vendas ---
    ws2 = wb.create_sheet("Pré-vendas")
    headers2 = ["Data Referência", "Código Unidade", "Nome", "País/Região", "Vendas Total"] + \
        [_pre_venda_label(k) for k in pv_keys] + ["ID do E-mail (UID)"]
    _write_header(ws2, headers2)

    for row in pv_rows:
        data_ref, codigo, nome, regiao, vendas_total, detalhe, msg_id = row
        detalhe = detalhe or {}
        line = [data_ref, codigo, nome, regiao, vendas_total] + [detalhe.get(k, 0) for k in pv_keys] + [msg_id]
        ws2.append(line)

    _autosize(ws2)
    ws2.freeze_panes = "C2"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path, len(main_rows), len(pv_rows)
